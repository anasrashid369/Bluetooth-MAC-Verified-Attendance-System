import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, datetime
import threading
import os

EXCEL_FILE = "attendance_register.xlsx"
STUDENTS_SHEET = "Students"
ATTENDANCE_SHEET = "Attendance"

lock = threading.Lock()

def _get_styles():
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    
    present_fill = PatternFill("solid", start_color="E2EFDA")
    absent_fill = PatternFill("solid", start_color="FFE0E0")
    present_font = Font(name="Arial", bold=True, size=10, color="375623")
    absent_font = Font(name="Arial", bold=True, size=10, color="9C0006")
    normal_font = Font(name="Arial", size=10)
    
    return {
        "border": border,
        "center": center,
        "left": left,
        "present_fill": present_fill,
        "absent_fill": absent_fill,
        "present_font": present_font,
        "absent_font": absent_font,
        "normal_font": normal_font
    }

def load_students():
    with lock:
        if not os.path.exists(EXCEL_FILE):
            return []
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[STUDENTS_SHEET]
        students = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0]: # Roll No exists
                students.append({
                    "roll_no": str(row[0]).strip(),
                    "name": str(row[1]).strip() if row[1] else "",
                    "mac": str(row[2]).strip() if row[2] else None,
                    "device_type": str(row[3]).strip() if row[3] else "",
                    "registration_date": str(row[4]).strip() if row[4] else "",
                    "status": str(row[5]).strip() if row[5] else "Not Registered"
                })
        return students

def student_already_registered(mac):
    students = load_students()
    for s in students:
        if s["mac"] and s["mac"].upper() == mac.upper():
            return s["name"]
    return None

def save_mac_registration(roll_no, mac, device_name):
    with lock:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[STUDENTS_SHEET]
        styles = _get_styles()
        
        target_row = None
        for row in ws.iter_rows(min_row=3):
            if str(row[0].value).strip() == str(roll_no).strip():
                target_row = row[0].row
                break
        
        if target_row:
            ws.cell(row=target_row, column=3).value = mac.upper()
            ws.cell(row=target_row, column=4).value = device_name
            ws.cell(row=target_row, column=5).value = str(date.today())
            ws.cell(row=target_row, column=6).value = "Registered"
            
            for col in range(1, 7):
                cell = ws.cell(row=target_row, column=col)
                cell.border = styles["border"]
                cell.alignment = styles["center"]
                if col == 6:
                    cell.fill = styles["present_fill"]
                    cell.font = styles["present_font"]
                else:
                    cell.font = styles["normal_font"]
            
            wb.save(EXCEL_FILE)
            return True
        return False

def get_next_attendance_row(ws):
    row = 3
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    return row

def mark_present(roll_no, mac, rssi, time_str):
    with lock:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_students = wb[STUDENTS_SHEET]
        ws_attend = wb[ATTENDANCE_SHEET]
        styles = _get_styles()
        
        student_name = ""
        for row in ws_students.iter_rows(min_row=3, values_only=True):
            if str(row[0]).strip() == str(roll_no).strip():
                student_name = row[1]
                break
        
        row_idx = get_next_attendance_row(ws_attend)
        values = [roll_no, student_name, mac, str(date.today()), time_str, f"{rssi} dBm", "PRESENT"]
        
        for col, val in enumerate(values, 1):
            cell = ws_attend.cell(row=row_idx, column=col, value=val)
            cell.border = styles["border"]
            cell.alignment = styles["center"] if col != 2 else styles["left"]
            if col == 7:
                cell.fill = styles["present_fill"]
                cell.font = styles["present_font"]
            else:
                cell.font = styles["normal_font"]
        
        wb.save(EXCEL_FILE)

def mark_absent(roll_no):
    with lock:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_students = wb[STUDENTS_SHEET]
        ws_attend = wb[ATTENDANCE_SHEET]
        styles = _get_styles()
        
        student_name = ""
        mac = ""
        for row in ws_students.iter_rows(min_row=3, values_only=True):
            if str(row[0]).strip() == str(roll_no).strip():
                student_name = row[1]
                mac = row[2]
                break
        
        row_idx = get_next_attendance_row(ws_attend)
        values = [roll_no, student_name, mac, str(date.today()), "-", "-", "ABSENT"]
        
        for col, val in enumerate(values, 1):
            cell = ws_attend.cell(row=row_idx, column=col, value=val)
            cell.border = styles["border"]
            cell.alignment = styles["center"] if col != 2 else styles["left"]
            if col == 7:
                cell.fill = styles["absent_fill"]
                cell.font = styles["absent_font"]
            else:
                cell.font = styles["normal_font"]
        
        wb.save(EXCEL_FILE)

def get_attendance_summary(target_date=None):
    if target_date is None:
        target_date = str(date.today())
    
    with lock:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb[ATTENDANCE_SHEET]
        present = 0
        absent = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if str(row[3]) == target_date:
                if row[6] == "PRESENT":
                    present += 1
                elif row[6] == "ABSENT":
                    absent += 1
        return {"present": present, "absent": absent}
